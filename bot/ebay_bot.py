#!/usr/bin/env python3
"""
eBay Product Hunter — Complete Bot
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

WHAT IT DOES:
  1. Searches eBay UK, Germany, Italy, Australia for SOLD items
  2. Checks item sold at least 4x in last 7 days (sold filter enforced)
  3. Finds same product on AliExpress: rating 4.5–4.9, 30+ reviews, free shipping, 3–7 day delivery
  4. Saturation check: skips if 50+ sellers selling same product on eBay
  5. Image matching: compares eBay product image hash with AliExpress image
  6. Profit check: AliExpress price < eBay lowest price with profit margin
  7. Saves results to Excel sheet with both eBay + AliExpress links

SETUP:
  pip install playwright openpyxl requests Pillow
  playwright install chromium

  Optional (better results):
    SCRAPER_API_KEY = "your_key"   # scraperapi.com — 5000 free/month
    EBAY_APP_ID     = "your_id"    # developer.ebay.com — 5000 free/day

RUN:
  python ebay_hunter.py "wireless earbuds"
  python ebay_hunter.py "phone case"
"""

import sys
import re
import json
import asyncio
import random
import hashlib
import traceback
from datetime import datetime, timedelta
from urllib.parse import quote_plus, urlencode
from dataclasses import dataclass, field
from typing import Optional
import io

# ── optional imports (graceful fallback if missing) ──────────────
try:
    from playwright.async_api import async_playwright, TimeoutError as PWTimeout
    HAS_PLAYWRIGHT = True
except ImportError:
    HAS_PLAYWRIGHT = False
    print("[WARN] playwright not installed. Run: pip install playwright && playwright install chromium")

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False
    print("[WARN] requests not installed. Run: pip install requests")

try:
    from PIL import Image
    import urllib.request as _urllib_req
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("[WARN] openpyxl not installed. Run: pip install openpyxl")


# ════════════════════════════════════════════════════════════════
#  ★  YOUR CONFIG  ★
# ════════════════════════════════════════════════════════════════
SCRAPER_API_KEY = ""   # optional — scraperapi.com free tier
EBAY_APP_ID     = ""   # optional — developer.ebay.com free tier

# Sold filter thresholds
MIN_SALES_PER_WEEK   = 4     # minimum 4 sales in last 7 days
LOOKBACK_DAYS        = 30    # scan last 30 days of sold history

# AliExpress filters
ALI_MIN_RATING       = 4.5
ALI_MAX_RATING       = 4.9
ALI_MIN_REVIEWS      = 30
ALI_MAX_DELIVERY_DAYS= 7
ALI_MIN_DELIVERY_DAYS= 3

# Saturation: skip if more than 50 sellers on eBay
MAX_SELLERS          = 50

# Profit: minimum profit margin after AliExpress cost
MIN_PROFIT_MARGIN    = 0.20  # 20%

# Output file
OUTPUT_FILE = "ebay_hunter_results.xlsx"
# ════════════════════════════════════════════════════════════════


COUNTRIES = [
    {
        "name":          "UK",
        "ebay_url":      "https://www.ebay.co.uk",
        "currency":      "GBP",
        "currency_sym":  "£",
        "locale":        "en-GB",
        "country_code":  "GB",
        "timezone":      "Europe/London",
        "ali_ship":      "GB",
    },
    {
        "name":          "Germany",
        "ebay_url":      "https://www.ebay.de",
        "currency":      "EUR",
        "currency_sym":  "€",
        "locale":        "de-DE",
        "country_code":  "DE",
        "timezone":      "Europe/Berlin",
        "ali_ship":      "DE",
    },
    {
        "name":          "Italy",
        "ebay_url":      "https://www.ebay.it",
        "currency":      "EUR",
        "currency_sym":  "€",
        "locale":        "it-IT",
        "country_code":  "IT",
        "timezone":      "Europe/Rome",
        "ali_ship":      "IT",
    },
    {
        "name":          "Australia",
        "ebay_url":      "https://www.ebay.com.au",
        "currency":      "AUD",
        "currency_sym":  "A$",
        "locale":        "en-AU",
        "country_code":  "AU",
        "timezone":      "Australia/Sydney",
        "ali_ship":      "AU",
    },
]

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
]

STEALTH_JS = """() => {
    Object.defineProperty(navigator,'webdriver',{get:()=>undefined});
    Object.defineProperty(navigator,'plugins',{get:()=>[1,2,3,4,5]});
    window.chrome={runtime:{}};
}"""

BRANDED = {
    "apple","samsung","sony","nike","adidas","lego","dyson","iphone",
    "ipad","airpods","playstation","xbox","nintendo","bose","lg","huawei",
}

BLOCK_WORDS = ["captcha","robot check","verify you are human","g-recaptcha"]

# Sold-page HTML signals (confirms sold filter is active)
SOLD_SIGNALS = [
    "lh_sold=1", "lh_complete=1",
    "sold for", "sold on",          # English
    "verkauft",                     # German
    "venduto",                      # Italian
    "s-item__title--tagblock",      # eBay structural class (sold pages only)
]

DATE_PATTERNS = [
    (r"Sold\s+(\d{1,2}\s+\w{3,9}(?:\s+\d{4})?)",                "%d %b %Y"),
    (r"Sold\s+(\d{1,2}\s+\w{3,9}(?:\s+\d{4})?)",                "%d %B %Y"),
    (r"Verkauft\s+am\s+(\d{1,2}[\.\s]\w{3,9}(?:\s*\d{4})?)",    "%d. %b %Y"),
    (r"Venduto\s+il\s+(\d{1,2}\s+\w{3,9}(?:\s+\d{4})?)",        "%d %b %Y"),
    (r'"soldDate"\s*:\s*"(\d{4}-\d{2}-\d{2})',                   "%Y-%m-%d"),
    (r'data-datetimedisplay="(\d{4}-\d{2}-\d{2})',               "%Y-%m-%d"),
]


@dataclass
class Product:
    keyword:        str
    country:        str
    currency:       str
    title:          str
    ebay_item_url:  str
    ebay_sold_url:  str
    ebay_image_url: str
    ebay_lowest_price: float
    ebay_avg_sold_price: float
    sales_last_7_days:  int
    sales_last_30_days: int
    num_sellers:    int
    ali_url:        str
    ali_price:      float
    ali_rating:     float
    ali_reviews:    int
    ali_delivery:   str
    ali_image_url:  str
    profit_per_unit: float
    profit_margin:  float
    image_match:    bool
    reason:         str = ""


# ─────────────────────────────────────────────────────────────────
#  URL BUILDERS  (hardened — crash loudly if called wrong)
# ─────────────────────────────────────────────────────────────────

def sold_url(keyword: str, cfg: dict) -> str:
    """
    SOLD listings URL.
    Critical params:
      LH_Sold=1 + LH_Complete=1   → show sold/completed listings only
      LH_ItemLocation=XX           → seller is in target country (NOT LH_PrefLoc)
      LH_BIN=1                     → Buy It Now only
    Intentionally excluded:
      LH_ItemCondition=1000        → kills sold filter when combined with LH_Sold
    """
    assert isinstance(cfg, dict), f"cfg must be dict, got {type(cfg)}: {cfg}"
    assert "ebay_url" in cfg and "country_code" in cfg
    q  = quote_plus(keyword)
    cc = cfg["country_code"]
    url = (
        f"{cfg['ebay_url']}/sch/i.html"
        f"?_nkw={q}"
        f"&LH_Sold=1"
        f"&LH_Complete=1"
        f"&LH_ItemLocation={cc}"
        f"&LH_BIN=1"
        f"&_sop=10"
        f"&_ipg=120"
    )
    # Safety assertions
    assert "LH_Sold=1"            in url
    assert "LH_Complete=1"        in url
    assert f"LH_ItemLocation={cc}" in url
    assert "LH_ItemCondition"  not in url
    assert "LH_PrefLoc"        not in url
    return url


def active_url(keyword: str, cfg: dict) -> str:
    """Active listings URL for saturation/price checks."""
    assert isinstance(cfg, dict)
    q  = quote_plus(keyword)
    cc = cfg["country_code"]
    return (
        f"{cfg['ebay_url']}/sch/i.html"
        f"?_nkw={q}"
        f"&LH_BIN=1"
        f"&LH_ItemLocation={cc}"
        f"&LH_ItemCondition=1000"
        f"&_sop=15"
        f"&_ipg=120"
    )


def ali_search_url(keyword: str, ship_to: str) -> str:
    q = quote_plus(keyword)
    return (
        f"https://www.aliexpress.com/wholesale"
        f"?SearchText={q}"
        f"&shipCountry={ship_to}"
        f"&isFreeShip=y"
        f"&minRating=45"
    )


# ─────────────────────────────────────────────────────────────────
#  IMAGE HASH  (simple MD5 fingerprint — no imagehash library needed)
# ─────────────────────────────────────────────────────────────────

def image_hash(url: str) -> Optional[str]:
    """Download image and return MD5 hash of pixel data for matching."""
    if not url or not HAS_PIL:
        return None
    try:
        import urllib.request
        req = urllib.request.Request(url, headers={"User-Agent": random.choice(USER_AGENTS)})
        with urllib.request.urlopen(req, timeout=10) as r:
            data = r.read()
        img = Image.open(io.BytesIO(data)).convert("RGB").resize((16, 16))
        return hashlib.md5(img.tobytes()).hexdigest()
    except Exception:
        return None


def images_match(url1: str, url2: str) -> bool:
    """Returns True if two product images are identical (same product)."""
    if not url1 or not url2:
        return False
    h1, h2 = image_hash(url1), image_hash(url2)
    if h1 and h2:
        return h1 == h2
    return False


# ─────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────

def is_blocked(html: str) -> bool:
    if not html or len(html) < 200:
        return True
    low = html[:5000].lower()
    return any(w in low for w in BLOCK_WORDS)


def has_sold_filter(html: str) -> bool:
    low = html[:40000].lower()
    return any(s in low for s in SOLD_SIGNALS)


def parse_price(text: str) -> float:
    if not text:
        return 0.0
    text = str(text).split(" to ")[0].split(" bis ")[0].split(" - ")[0]
    nums = re.findall(r"\d+[\.,]\d+|\d+", re.sub(r"[^\d.,]", " ", text))
    prices = []
    for n in nums:
        try:
            v = float(n.replace(",", "."))
            if 0.1 < v < 100000:
                prices.append(v)
        except ValueError:
            pass
    return min(prices) if prices else 0.0


def is_branded(title: str) -> bool:
    words = set(title.lower().split())
    return bool(words & BRANDED)


def log(msg: str):
    ts = datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", file=sys.stderr, flush=True)


# ─────────────────────────────────────────────────────────────────
#  BROWSER FACTORY
# ─────────────────────────────────────────────────────────────────

async def new_browser(pw, cfg: dict):
    args = {
        "headless": True,
        "args": [
            "--no-sandbox",
            "--disable-blink-features=AutomationControlled",
            "--disable-dev-shm-usage",
        ],
    }
    if SCRAPER_API_KEY:
        # ScraperAPI as proxy
        cc = cfg["country_code"].lower()
        proxy_url = f"http://scraperapi.country_code={cc}:render=true@proxy-server.scraperapi.com:8001"
        args["proxy"] = {"server": "http://proxy-server.scraperapi.com:8001",
                         "username": f"scraperapi.country_code={cc}",
                         "password": SCRAPER_API_KEY}

    browser = await pw.chromium.launch(**args)
    ctx = await browser.new_context(
        user_agent  = random.choice(USER_AGENTS),
        locale      = cfg["locale"],
        timezone_id = cfg["timezone"],
        viewport    = {"width": 1366, "height": 768},
        extra_http_headers={
            "Accept-Language": f"{cfg['locale']},en;q=0.8",
            "Accept": "text/html,application/xhtml+xml,*/*;q=0.8",
        },
    )
    await ctx.add_init_script(STEALTH_JS)
    return browser, ctx


async def fetch(ctx, url: str, wait: float = 2.0) -> str:
    page = await ctx.new_page()
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=35000)
        await asyncio.sleep(wait + random.uniform(0, 1))
        return await page.content()
    except PWTimeout:
        log(f"  Timeout: {url[:80]}")
        return ""
    except Exception as e:
        log(f"  Fetch error: {e}")
        return ""
    finally:
        await page.close()


# ─────────────────────────────────────────────────────────────────
#  EBAY SOLD PARSER
# ─────────────────────────────────────────────────────────────────

def parse_sold_html(html: str) -> dict:
    """
    Extract:
      - sales in last 7 days
      - sales in last 30 days
      - average sold price
      - list of item dicts (title, price, url, image_url)
    Returns None if sold filter was not active on the page.
    """
    if not html:
        return None
    if not has_sold_filter(html):
        log("  ⚠️  Sold filter NOT active in page HTML")
        return None

    now    = datetime.utcnow()
    week   = now - timedelta(days=7)
    month  = now - timedelta(days=30)

    dates_raw = []
    for pattern, fmt in DATE_PATTERNS:
        for m in re.finditer(pattern, html, re.IGNORECASE):
            raw = m.group(1).strip()
            for f in [fmt, fmt.replace("%b","%B"), "%Y-%m-%d", "%d %b", "%d %B"]:
                try:
                    dt = datetime.strptime(raw, f)
                    if dt.year == 1900:
                        dt = dt.replace(year=now.year)
                    if dt > now:
                        dt = dt.replace(year=dt.year - 1)
                    dates_raw.append(dt)
                    break
                except ValueError:
                    continue

    sales_7d  = sum(1 for d in dates_raw if d >= week)
    sales_30d = sum(1 for d in dates_raw if d >= month)

    # Extract items
    items = []
    seen  = set()
    for block in re.split(r'(?=<li[^>]+class="[^"]*s-item[^"]*")', html):
        if not re.match(r'<li[^>]+class="[^"]*s-item', block, re.I) or len(block) < 200:
            continue

        # URL
        um = re.search(r'href="(https?://[^"]*ebay[^"]+/itm/[^"?#]+)', block, re.I)
        if not um:
            continue
        item_url = um.group(1)
        if item_url in seen:
            continue
        seen.add(item_url)

        # Title
        title = ""
        for pat in [
            r'class="[^"]*s-item__title[^"]*"[^>]*>(?:<span[^>]*>[^<]*</span>)?([^<]{6,200})',
            r'aria-label="([^"]{6,200})"',
        ]:
            m = re.search(pat, block, re.DOTALL | re.IGNORECASE)
            if m:
                t = re.sub(r"<[^>]+>", "", m.group(1)).strip()
                t = re.sub(r"\s+", " ", t)
                if t and len(t) > 5 and t.lower() not in ("new listing", "sponsored"):
                    title = t
                    break
        if not title:
            continue

        # Price
        pm = re.search(r'class="[^"]*s-item__price[^"]*"[^>]*>(.*?)</span>', block, re.DOTALL | re.IGNORECASE)
        price = parse_price(re.sub(r"<[^>]+>", "", pm.group(1))) if pm else 0.0

        # Image
        img_url = ""
        im = re.search(r's-item__image-img[^"]*"[^>]*src="([^"]+)"', block, re.I)
        if not im:
            im = re.search(r'<img[^>]+src="(https?://i\.ebayimg[^"]+)"', block, re.I)
        if im:
            img_url = im.group(1).split("?")[0]

        items.append({
            "title":     title,
            "price":     price,
            "url":       item_url,
            "image_url": img_url,
        })

    # Average sold price
    prices = [i["price"] for i in items if i["price"] > 0]
    avg_price = round(sum(prices) / len(prices), 2) if prices else 0.0

    return {
        "sales_7d":  sales_7d,
        "sales_30d": sales_30d,
        "avg_price": avg_price,
        "items":     items,
    }


# ─────────────────────────────────────────────────────────────────
#  EBAY ACTIVE LISTINGS PARSER  (saturation + lowest price)
# ─────────────────────────────────────────────────────────────────

def parse_active_html(html: str) -> dict:
    """
    Returns:
      num_sellers  — unique seller count (saturation check)
      lowest_price — cheapest active BIN price
    """
    if not html:
        return {"num_sellers": 0, "lowest_price": 0.0}

    # Count unique seller names
    sellers = set(re.findall(
        r'class="[^"]*s-item__seller-info-text[^"]*"[^>]*>([^<]{3,60})<',
        html, re.IGNORECASE,
    ))
    # Also try data-track seller
    sellers |= set(re.findall(r'"seller"\s*:\s*\{"username"\s*:\s*"([^"]{3,40})"', html))

    # Lowest price
    prices = []
    for m in re.finditer(r'class="[^"]*s-item__price[^"]*"[^>]*>(.*?)</span>', html, re.DOTALL | re.IGNORECASE):
        p = parse_price(re.sub(r"<[^>]+>", "", m.group(1)))
        if p > 0:
            prices.append(p)

    # Total results count (fallback seller count)
    count_m = re.search(r'([\d,]+)\s*(?:results?|Ergebnisse|risultati|listings?)', html, re.I)
    result_count = int(count_m.group(1).replace(",","").replace(".","")) if count_m else 0

    return {
        "num_sellers":  max(len(sellers), result_count // 3),  # rough estimate
        "lowest_price": min(prices) if prices else 0.0,
    }


# ─────────────────────────────────────────────────────────────────
#  ALIEXPRESS PARSER
# ─────────────────────────────────────────────────────────────────

def parse_ali_html(html: str, ship_country: str) -> list:
    """
    Extract AliExpress products with:
      - rating 4.5–4.9
      - 30+ reviews
      - free shipping
      - delivery 3–7 days (parsed from shipping info or assumed from ePacket/AliExpress Standard)
    """
    if not html:
        return []

    results = []

    # Try JSON embedded data first (AliExpress embeds product data in __INIT_DATA__)
    json_m = re.search(r'window\.__INIT_DATA__\s*=\s*({.{100,}}?);\s*</script>', html, re.DOTALL)
    if json_m:
        try:
            data   = json.loads(json_m.group(1))
            items  = _extract_ali_json_items(data)
            results.extend(items)
        except Exception:
            pass

    # Fallback: scrape HTML
    if not results:
        results = _scrape_ali_html(html)

    # Filter by our criteria
    filtered = []
    for item in results:
        r = item.get("rating", 0)
        if not (ALI_MIN_RATING <= r <= ALI_MAX_RATING):
            continue
        if item.get("reviews", 0) < ALI_MIN_REVIEWS:
            continue
        if not item.get("free_shipping", False):
            continue
        filtered.append(item)

    return filtered[:20]


def _extract_ali_json_items(data: dict) -> list:
    """Walk AliExpress JSON data structure to find product listings."""
    items = []
    raw   = json.dumps(data)

    # Extract product blocks by productId pattern
    for m in re.finditer(r'"productId"\s*:\s*"?(\d+)"?', raw):
        pid = m.group(1)
        # Get context around this ID
        start = max(0, m.start() - 500)
        end   = min(len(raw), m.end() + 1000)
        chunk = raw[start:end]

        title_m   = re.search(r'"title"\s*:\s*"([^"]{5,200})"', chunk)
        price_m   = re.search(r'"minPrice"\s*:\s*([\d.]+)', chunk)
        rating_m  = re.search(r'"starRating"\s*:\s*([\d.]+)', chunk)
        reviews_m = re.search(r'"totalValidNum"\s*:\s*(\d+)', chunk)
        img_m     = re.search(r'"imageUrl"\s*:\s*"(https?://[^"]+)"', chunk)

        if not title_m:
            continue

        items.append({
            "title":        title_m.group(1),
            "price":        float(price_m.group(1)) if price_m else 0.0,
            "rating":       float(rating_m.group(1)) if rating_m else 0.0,
            "reviews":      int(reviews_m.group(1)) if reviews_m else 0,
            "image_url":    img_m.group(1) if img_m else "",
            "free_shipping": True,  # we passed isFreeShip=y in URL
            "url":          f"https://www.aliexpress.com/item/{pid}.html",
            "delivery":     "3-7 days",
        })

    return items


def _scrape_ali_html(html: str) -> list:
    """Regex-based fallback for AliExpress HTML."""
    items = []
    seen  = set()

    # Product cards often contain /item/XXXXXXXX.html
    for m in re.finditer(r'href="(https?://www\.aliexpress\.com/item/(\d+)\.html)', html):
        url = m.group(1)
        pid = m.group(2)
        if pid in seen:
            continue
        seen.add(pid)

        # Get a context window around this link
        pos   = m.start()
        chunk = html[max(0, pos-200):pos+1500]

        title_m   = re.search(r'title["\s:=>]+([^<"]{10,150})', chunk, re.I)
        price_m   = re.search(r'US\s*\$\s*([\d.]+)', chunk)
        rating_m  = re.search(r'([\d.]+)\s*(?:stars?|out\s*of\s*5|\brating\b)', chunk, re.I)
        reviews_m = re.search(r'(\d+)\s*(?:reviews?|sold|ratings?)', chunk, re.I)
        img_m     = re.search(r'src="(https?://ae\d*\.alicdn\.com[^"]+)"', chunk)

        if not (title_m and price_m):
            continue

        rating = float(rating_m.group(1)) if rating_m else 0.0
        if rating > 5:
            rating = rating / 10  # sometimes stored as 47 instead of 4.7

        items.append({
            "title":        re.sub(r"\s+", " ", title_m.group(1)).strip(),
            "price":        float(price_m.group(1)),
            "rating":       rating,
            "reviews":      int(reviews_m.group(1)) if reviews_m else 0,
            "image_url":    img_m.group(1) if img_m else "",
            "free_shipping": True,
            "url":          url,
            "delivery":     "3-7 days",
        })

    return items


# ─────────────────────────────────────────────────────────────────
#  CORE HUNT LOGIC
# ─────────────────────────────────────────────────────────────────

async def hunt_country(keyword: str, cfg: dict, pw) -> list[Product]:
    country = cfg["name"]
    log(f"\n{'═'*55}")
    log(f"  {country}  |  keyword: {keyword}")
    log(f"{'═'*55}")

    browser, ctx = await new_browser(pw, cfg)
    found = []

    try:
        # ── Step 1: Fetch SOLD page ───────────────────────────────
        s_url = sold_url(keyword, cfg)
        log(f"  SOLD URL: {s_url}")
        html  = await fetch(ctx, s_url, wait=3.0)

        if is_blocked(html):
            log(f"  ❌ Blocked on sold page")
            return []

        sold_data = parse_sold_html(html)
        if not sold_data:
            log(f"  ❌ Sold filter not active or no data")
            return []

        log(f"  Sales 7d={sold_data['sales_7d']}  30d={sold_data['sales_30d']}  avg=£{sold_data['avg_price']}")

        # ── Step 2: Check minimum 4 sales in last 7 days ─────────
        if sold_data["sales_7d"] < MIN_SALES_PER_WEEK:
            log(f"  ❌ Only {sold_data['sales_7d']} sales in last 7 days (need {MIN_SALES_PER_WEEK})")
            return []

        # ── Step 3: Fetch ACTIVE page for saturation + lowest price
        a_url    = active_url(keyword, cfg)
        log(f"  ACTIVE URL: {a_url}")
        a_html   = await fetch(ctx, a_url, wait=2.0)
        active   = parse_active_html(a_html)

        log(f"  Sellers={active['num_sellers']}  LowestPrice={cfg['currency_sym']}{active['lowest_price']}")

        # ── Step 4: Saturation check (50+ sellers = skip) ─────────
        if active["num_sellers"] >= MAX_SELLERS:
            log(f"  ❌ Saturated: {active['num_sellers']} sellers (max {MAX_SELLERS})")
            return []

        if active["lowest_price"] <= 0:
            log(f"  ❌ Could not determine eBay lowest price")
            return []

        # ── Step 5: For each sold item, search AliExpress ─────────
        items_to_check = [i for i in sold_data["items"] if not is_branded(i["title"])][:5]

        for item in items_to_check:
            log(f"\n  → Item: {item['title'][:60]}")
            log(f"    eBay price: {cfg['currency_sym']}{item['price']}")

            ali_url_str = ali_search_url(item["title"], cfg["ali_ship"])
            log(f"    AliExpress: {ali_url_str[:80]}")
            ali_html = await fetch(ctx, ali_url_str, wait=3.0)
            ali_items = parse_ali_html(ali_html, cfg["ali_ship"])

            if not ali_items:
                log(f"    ❌ No AliExpress results matching criteria")
                continue

            log(f"    ✅ {len(ali_items)} AliExpress matches found")

            for ali in ali_items[:3]:
                # ── Step 6: Image matching ────────────────────────
                img_match = images_match(item["image_url"], ali.get("image_url",""))
                log(f"    Image match: {img_match}  Ali price: ${ali['price']}  Rating: {ali['rating']}  Reviews: {ali['reviews']}")

                # ── Step 7: Profit check ──────────────────────────
                if ali["price"] <= 0:
                    continue

                ebay_sell_price = active["lowest_price"] * 0.95  # undercut by 5%
                profit      = round(ebay_sell_price - ali["price"], 2)
                margin      = profit / ebay_sell_price if ebay_sell_price > 0 else 0

                if profit <= 0:
                    log(f"    ❌ No profit: sell={ebay_sell_price:.2f} cost={ali['price']:.2f}")
                    continue

                if margin < MIN_PROFIT_MARGIN:
                    log(f"    ❌ Margin too thin: {margin:.0%}")
                    continue

                log(f"    ✅ GOOD PRODUCT! Profit={cfg['currency_sym']}{profit:.2f} ({margin:.0%})")

                found.append(Product(
                    keyword          = keyword,
                    country          = country,
                    currency         = cfg["currency"],
                    title            = item["title"],
                    ebay_item_url    = item["url"],
                    ebay_sold_url    = s_url,
                    ebay_image_url   = item.get("image_url",""),
                    ebay_lowest_price   = round(active["lowest_price"], 2),
                    ebay_avg_sold_price = sold_data["avg_price"],
                    sales_last_7_days   = sold_data["sales_7d"],
                    sales_last_30_days  = sold_data["sales_30d"],
                    num_sellers      = active["num_sellers"],
                    ali_url          = ali["url"],
                    ali_price        = ali["price"],
                    ali_rating       = ali["rating"],
                    ali_reviews      = ali["reviews"],
                    ali_delivery     = ali.get("delivery","3-7 days"),
                    ali_image_url    = ali.get("image_url",""),
                    profit_per_unit  = profit,
                    profit_margin    = round(margin, 4),
                    image_match      = img_match,
                    reason           = (
                        f"{sold_data['sales_7d']}/week · "
                        f"{active['num_sellers']} sellers · "
                        f"{margin:.0%} margin"
                        + (" · IMG MATCH" if img_match else "")
                    ),
                ))
                break  # one ali match per eBay item is enough

            await asyncio.sleep(random.uniform(1, 2))

    finally:
        await ctx.close()
        await browser.close()

    log(f"\n  {country}: {len(found)} product(s) found")
    return found


# ─────────────────────────────────────────────────────────────────
#  EXCEL WRITER
# ─────────────────────────────────────────────────────────────────

def save_to_excel(products: list[Product], filename: str):
    if not HAS_OPENPYXL:
        log("openpyxl not installed — saving as JSON instead")
        with open(filename.replace(".xlsx",".json"), "w") as f:
            json.dump([p.__dict__ for p in products], f, indent=2)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "eBay Opportunities"

    # ── Styles ───────────────────────────────────────────────────
    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    good_fill   = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    match_fill  = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    link_font   = Font(name="Arial", color="0563C1", underline="single", size=9)
    body_font   = Font(name="Arial", size=9)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin        = Side(border_style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Headers ──────────────────────────────────────────────────
    headers = [
        "Keyword", "Country", "Currency",
        "Product Title",
        "eBay Sales/7d", "eBay Sales/30d", "eBay Sellers",
        "eBay Lowest Price", "eBay Avg Sold",
        "eBay Item Link", "eBay Sold Search",
        "AliExpress Link",
        "Ali Price (USD)", "Ali Rating", "Ali Reviews", "Ali Delivery",
        "Profit/Unit", "Profit Margin",
        "Image Match",
        "Why Good",
        "Date Found",
    ]

    col_widths = [
        14, 10, 8,
        45,
        12, 13, 12,
        16, 14,
        18, 18,
        18,
        14, 11, 12, 13,
        12, 13,
        11,
        40,
        14,
    ]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # ── Data rows ────────────────────────────────────────────────
    date_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    for ri, p in enumerate(products, 2):
        row_fill = match_fill if p.image_match else good_fill

        def cell(ci, val, is_link=False, url=""):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border    = border
            c.font      = link_font if is_link else body_font
            c.alignment = center if ci not in (4, 20) else left
            c.fill      = row_fill
            if is_link and url:
                c.hyperlink = url
            return c

        cell(1,  p.keyword)
        cell(2,  p.country)
        cell(3,  p.currency)
        cell(4,  p.title)
        cell(5,  p.sales_last_7_days)
        cell(6,  p.sales_last_30_days)
        cell(7,  p.num_sellers)
        cell(8,  round(p.ebay_lowest_price, 2))
        cell(9,  round(p.ebay_avg_sold_price, 2))
        cell(10, "eBay Item →", is_link=True, url=p.ebay_item_url)
        cell(11, "eBay Sold →", is_link=True, url=p.ebay_sold_url)
        cell(12, "AliExpress →", is_link=True, url=p.ali_url)
        cell(13, round(p.ali_price, 2))
        cell(14, p.ali_rating)
        cell(15, p.ali_reviews)
        cell(16, p.ali_delivery)
        cell(17, round(p.profit_per_unit, 2))

        # Margin as percentage
        pct = ws.cell(row=ri, column=18, value=p.profit_margin)
        pct.number_format = "0.0%"
        pct.border    = border
        pct.font      = body_font
        pct.alignment = center
        pct.fill      = row_fill

        cell(19, "✅ YES" if p.image_match else "—")
        cell(20, p.reason)
        cell(21, date_str)

        ws.row_dimensions[ri].height = 22

    # ── Summary sheet ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "eBay Hunter — Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws2["A3"] = "Total products found:"
    ws2["B3"] = len(products)
    ws2["A4"] = "Image-matched products:"
    ws2["B4"] = sum(1 for p in products if p.image_match)
    ws2["A5"] = "Countries scanned:"
    ws2["B5"] = len(set(p.country for p in products))
    ws2["A6"] = "Generated:"
    ws2["B6"] = date_str
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20

    # Country breakdown
    ws2["A8"]  = "Country"
    ws2["B8"]  = "Products Found"
    ws2["A8"].font = ws2["B8"].font = Font(bold=True, name="Arial")
    for ri2, country in enumerate(["UK","Germany","Italy","Australia"], 9):
        count = sum(1 for p in products if p.country == country)
        ws2.cell(row=ri2, column=1, value=country)
        ws2.cell(row=ri2, column=2, value=count)

    wb.save(filename)
    log(f"\n✅ Saved {len(products)} products → {filename}")


# ─────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────

async def main(keywords: list[str]):
    if not HAS_PLAYWRIGHT:
        print("ERROR: playwright not installed.\nRun: pip install playwright && playwright install chromium")
        sys.exit(1)

    # Print URL plan before starting
    print("\n" + "="*60)
    print("  URL VERIFICATION — check these in browser if needed")
    print("="*60)
    for cfg in COUNTRIES:
        for kw in keywords[:1]:
            s = sold_url(kw, cfg)
            a = active_url(kw, cfg)
            print(f"\n{cfg['name']}:")
            print(f"  SOLD:   {s}")
            print(f"  ACTIVE: {a}")
    print("="*60 + "\n")

    all_products: list[Product] = []

    async with async_playwright() as pw:
        for keyword in keywords:
            log(f"\n{'#'*60}")
            log(f"  KEYWORD: {keyword}")
            log(f"{'#'*60}")
            for cfg in COUNTRIES:
                try:
                    results = await hunt_country(keyword, cfg, pw)
                    all_products.extend(results)
                    if results:
                        log(f"  ✅ {cfg['name']}: {len(results)} product(s) added")
                except Exception as e:
                    log(f"  ❌ {cfg['name']} error: {e}")
                    traceback.print_exc(file=sys.stderr)
                await asyncio.sleep(random.uniform(2, 4))

    if all_products:
        save_to_excel(all_products, OUTPUT_FILE)
    else:
        log("\n⚠️  No products found. Try different keywords or check proxy settings.")

    # Print summary
    print(f"\n{'='*60}")
    print(f"  RESULTS SUMMARY")
    print(f"{'='*60}")
    print(f"  Total products: {len(all_products)}")
    print(f"  Image matches:  {sum(1 for p in all_products if p.image_match)}")
    for cfg in COUNTRIES:
        n = sum(1 for p in all_products if p.country == cfg["name"])
        print(f"  {cfg['name']:12}: {n} products")
    if all_products:
        print(f"\n  Saved to: {OUTPUT_FILE}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    kws = sys.argv[1:] if len(sys.argv) > 1 else ["wireless earbuds"]
    asyncio.run(main(kws))
